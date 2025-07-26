import sys
import os
import adsk.core, adsk.fusion, json
from .LogUtils import LogUtils

plugin_dir = os.path.dirname(os.path.abspath(__file__))
if plugin_dir not in sys.path:
    sys.path.insert(0, plugin_dir)

class ConfigUtils:

    @staticmethod
    def write_configs_to_excel(file_path: str, configs: list, parameters: list):
        """
        将配置写入Excel文件
        :param file_path: Excel文件路径
        :param configs: 配置列表
        :param parameters: 参数列表
        """
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "导出配置"
            
            # 构建带注释的表头
            headers = ['导出格式', '自定义名称']
            for param in parameters:
                param_name = param['name']
                comment = param.get('comment', '').strip()
                if comment:
                    # 如果注释包含换行符，只取第一行
                    comment = comment.split('\n')[0].strip()
                    # 限制注释长度，避免表头过长
                    if len(comment) > 25:
                        comment = comment[:22] + '...'
                    header = f"{param_name}\n({comment})"
                else:
                    header = param_name
                headers.append(header)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # 写入配置数据
            for row, config in enumerate(configs, 2):
                ws.cell(row=row, column=1, value=config.get('format', 'step'))
                ws.cell(row=row, column=2, value=config.get('name', ''))
                
                for col, param in enumerate(parameters, 3):
                    param_value = config.get('parameters', {}).get(param['name'], param['expression'])
                    ws.cell(row=row, column=col, value=param_value)
            
            # 调整列宽
            for col in range(1, len(headers) + 1):
                if col <= 2:  # 前两列（导出格式、自定义名称）
                    ws.column_dimensions[get_column_letter(col)].width = 15
                else:  # 参数列，需要更宽以容纳换行
                    ws.column_dimensions[get_column_letter(col)].width = 20
            
            wb.save(file_path)
            LogUtils.info(f'配置已保存到Excel文件: {file_path}')
            return True
        except Exception as e:
            LogUtils.error(f'保存Excel文件失败: {str(e)}')
            return False

    @staticmethod
    def read_configs_from_excel(file_path: str, parameters: list):
        """
        从Excel文件读取配置
        :param file_path: Excel文件路径
        :param parameters: 参数列表
        :return: 配置列表或None
        """
        try:
            from openpyxl import load_workbook
            
            if not os.path.exists(file_path):
                LogUtils.error(f'Excel文件不存在: {file_path}')
                return None
            
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 验证表头
            header_validation = ConfigUtils._validate_excel_headers(ws, parameters)
            if not header_validation['valid']:
                error_msg = f"Excel文件表头验证失败:\n{header_validation['error_msg']}"
                LogUtils.error(error_msg)
                # 显示错误信息给用户
                import adsk.core
                ui = adsk.core.Application.get().userInterface
                if ui:
                    ui.messageBox(error_msg)
                return None
            
            configs = []
            
            # 从第二行开始读取数据（第一行是表头）
            for row in range(2, ws.max_row + 1):
                format_val = ws.cell(row=row, column=1).value
                name_val = ws.cell(row=row, column=2).value
                
                if not format_val and not name_val:
                    continue  # 跳过空行
                
                config = {
                    'format': format_val or 'step',
                    'name': name_val or '',
                    'parameters': {}
                }
                
                # 读取参数值
                for col, param in enumerate(parameters, 3):
                    param_value = ws.cell(row=row, column=col).value
                    if param_value is not None:
                        config['parameters'][param['name']] = str(param_value)
                
                configs.append(config)
            
            LogUtils.info(f'从Excel文件读取了 {len(configs)} 个配置: {file_path}')
            return configs
        except Exception as e:
            LogUtils.error(f'读取Excel文件失败: {str(e)}')
            return None

    @staticmethod
    def _extract_param_name_from_header(header):
        """从表头中提取参数名"""
        if not header:
            return ''
        
        # 处理换行格式：参数名\n(注释)
        if '\n' in header:
            return header.split('\n')[0].strip()
        
        # 处理括号格式：参数名 (注释)
        if '(' in header and header.endswith(')'):
            return header.split('(')[0].strip()
        
        # 没有注释的普通参数名
        return header.strip()

    @staticmethod
    def _validate_excel_headers(ws, parameters):
        """验证Excel表头是否与当前参数匹配"""
        try:
            # 获取Excel表头
            excel_headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    excel_headers.append(str(cell_value).strip())
            
            # 检查基本表头
            if len(excel_headers) < 2:
                return {
                    'valid': False,
                    'error_msg': 'Excel文件表头不完整，至少需要"导出格式"和"自定义名称"两列'
                }
            
            if excel_headers[0] != '导出格式' or excel_headers[1] != '自定义名称':
                return {
                    'valid': False,
                    'error_msg': f'Excel文件表头格式错误，前两列应为"导出格式"和"自定义名称"，实际为"{excel_headers[0]}"和"{excel_headers[1]}"'
                }
            
            # 提取Excel中的参数名
            excel_param_names = []
            for header in excel_headers[2:]:  # 跳过前两列
                param_name = ConfigUtils._extract_param_name_from_header(header)
                if param_name:
                    excel_param_names.append(param_name)
            
            # 获取当前设计中的参数名
            current_param_names = [param['name'] for param in parameters]
            
            # 检查缺失的参数
            missing_params = [name for name in current_param_names if name not in excel_param_names]
            if missing_params:
                return {
                    'valid': False,
                    'error_msg': f'Excel文件中缺少以下参数列：\n{", ".join(missing_params)}\n\n请重新导出Excel模板以获取完整的参数列表。'
                }
            
            # 检查多余的参数（可选，给出警告）
            extra_params = [name for name in excel_param_names if name not in current_param_names]
            if extra_params:
                LogUtils.warn(f'Excel文件中包含当前设计中不存在的参数：{", ".join(extra_params)}')
            
            return {'valid': True, 'error_msg': ''}
            
        except Exception as e:
            return {
                'valid': False,
                'error_msg': f'验证表头时发生错误：{str(e)}'
            }

    @staticmethod
    def create_excel_template(file_path: str, parameters: list):
        """
        创建/补全Excel模板文件，保留原有数据，只补充缺失的参数列。
        :param file_path: Excel文件路径
        :param parameters: 参数列表
        """
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment
            import os
            
            # 构建带注释的表头
            param_names = []
            param_exprs = {}
            param_headers = []
            
            for param in parameters:
                param_name = param['name']
                param_names.append(param_name)
                param_exprs[param_name] = param['expression']
                
                # 构建带注释的表头
                comment = param.get('comment', '').strip()
                if comment:
                    # 如果注释包含换行符，只取第一行
                    comment = comment.split('\n')[0].strip()
                    # 限制注释长度，避免表头过长
                    if len(comment) > 25:
                        comment = comment[:22] + '...'
                    header = f"{param_name}\n({comment})"
                else:
                    header = param_name
                
                param_headers.append(header)
            
            headers = ['导出格式', '自定义名称'] + param_headers
            # 如果文件存在，读取原有数据
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
                # 读取原有表头
                old_headers = [cell.value for cell in ws[1]]
                # 计算缺失的参数列（需要从带注释的表头中提取原始参数名）
                old_param_names = []
                for header in old_headers:
                    if header not in ['导出格式', '自定义名称']:
                        # 提取原始参数名（去掉注释部分）
                        param_name = ConfigUtils._extract_param_name_from_header(header)
                        old_param_names.append(param_name)
                
                missing_params = [p for p in param_names if p not in old_param_names]
                # 如果没有缺失，直接返回
                if not missing_params:
                    LogUtils.info(f'Excel模板已存在且无缺失列: {file_path}')
                    return True
                # 补充缺失的表头
                missing_headers = []
                for param in missing_params:
                    # 为缺失的参数构建带注释的表头
                    comment = next((p.get('comment', '').strip() for p in parameters if p['name'] == param), '')
                    if comment:
                        comment = comment.split('\n')[0].strip()
                        if len(comment) > 30:
                            comment = comment[:27] + '...'
                        header = f"{param} ({comment})"
                    else:
                        header = param
                    missing_headers.append(header)
                
                new_headers = old_headers + missing_headers
                for i, header in enumerate(new_headers, 1):
                    cell = ws.cell(row=1, column=i, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                # 补充每行缺失的参数列
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    old_row_len = len(old_headers)
                    for idx, param in enumerate(missing_params, start=old_row_len+1):
                        row[0].parent.cell(row=row[0].row, column=idx, value=param_exprs.get(param, ''))
                # 调整列宽
                for col in range(1, len(new_headers) + 1):
                    if col <= 2:  # 前两列（导出格式、自定义名称）
                        ws.column_dimensions[get_column_letter(col)].width = 15
                    else:  # 参数列，需要更宽以容纳换行
                        ws.column_dimensions[get_column_letter(col)].width = 20
                wb.save(file_path)
                LogUtils.info(f'Excel模板已补全缺失列: {file_path}')
                return True
            # 文件不存在，创建新模板
            wb = Workbook()
            ws = wb.active
            ws.title = "导出配置"
            # 表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # 示例行
            example_row = ['step', 'default'] + [param_exprs[name] for name in param_names]
            for col, value in enumerate(example_row, 1):
                ws.cell(row=2, column=col, value=value)
            for col in range(1, len(headers) + 1):
                if col <= 2:  # 前两列（导出格式、自定义名称）
                    ws.column_dimensions[get_column_letter(col)].width = 15
                else:  # 参数列，需要更宽以容纳换行
                    ws.column_dimensions[get_column_letter(col)].width = 20
            wb.save(file_path)
            LogUtils.info(f'Excel模板已创建: {file_path}')
            return True
        except Exception as e:
            LogUtils.error(f'创建Excel模板失败: {str(e)}')
            return False 